import re
import openpyxl as op

# Создание файла для результата
# result_file = op.Workbook()
# result_file.create_sheet(title='result', index=0)
# result_sheet = result_file['result']

# Ввод имени файла для работы и проверка
name_of_source_file = input('Введите файл для проверки: ')

try:
    source_file = op.load_workbook(name_of_source_file)
except op.utils.exceptions.InvalidFileException:
    print('Не верное имя или формат файла. Поддерживаются файлы с расширениями .xlsx,.xlsm,.xltx,.xltm')
    exit()
except FileNotFoundError:
    print('Указанный файл не найден')
else:
    print('OK')

# Ввод названия листа для работы и проверка
name_of_sheet = input('Введите название листа: ')

try:
    sheet = source_file[name_of_sheet]
except KeyError:
    print('Листа с данным названием не существует')
    exit()
else:
    print('OK')

# Обозначение стобца для работы
column_char = input('Введите столбец с лейблами: ')
count = 2
j = column_char + str(count)

# Обозначение строки, по которой будут отфильтровываться нужные строки
string_for_search = input('Введите строку для поиска: ')
pattern = re.compile(re.escape(string_for_search))

# Проход по строкам файла, пока строки не пустые
while str(sheet[j].value) != "None":

    # Получаем результат поиска заданной строки
    result = pattern.search(str(sheet[j].value))

    # Если строка не найдена - вся строка в Excel файле удаляется
    # при этом счётчик уменьшается, так как на место обрабатываемой пришла следующая, которая не проверялась
    if result is None:
        # result_sheet['A' + str(count)] = str(sheet[j].value)
        sheet.delete_rows(count, 1)
        count -= 1

    # Прибавляем счётчик и задаём адрес следующей ячейки
    count += 1
    j = column_char + str(count)

# Сохранение скопированных лейблов в новом файле, закрытие файла
# result_file.save('result_file.xlsx')
# result_file.close()

# Сохранение изменений в файле, закрытие файла
source_file.save(name_of_source_file)
source_file.close()
